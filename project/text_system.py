#coding:UTF-8

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
import random, sys, importlib
action = input("1.只默写新背内容  2.复习旧内容并且默写新内容")
if action =="1":
    num = int(input("你需要默写多少个单词？"))
    path = input("文件地址：")
    wb = openpyxl.load_workbook(path)

    sheets = wb.get_sheet_names()
    print(sheets)

    for i in range(len(sheets)):
        sheet = wb.get_sheet_by_name(sheets[i])

        print('\n\n'+str(i+1)+'sheet: '+sheet.title+'->>>')
        data_list = []
        for rownum in random.sample(range(1,100),num):
            for columnnum in range(1, sheet.max_column+1):
                data_list.append(sheet.cell(row = rownum, column = columnnum).value)

    for i in data_list:
        print(i)
    a = []
    for i in range(len(data_list)):
        if i % 2 ==0:
            a.append(data_list[i])
    for i in a:
        print(i)

    wbl = Workbook()
    new_ws = wbl.create_sheet(title='test')
    key_ws = wbl.create_sheet(title='test_key')
    i = 1
    k = 1
    for data_1 in range(len(data_list)):
        key_ws.cell(column=k, row=i, value=data_list[data_l])
        if k % 2 == 1:
            k = 2
        else:
            k = 1
            i += 1
            
    i = 1
    for data_m in range(len(a)):
        new_ws.cell(column=1, row=i, value=a[data_m])
        i += 1
    wbl.save(filename = 'new_file.xlsx')
elif action == "2":
    num = int(input("你需要默写多少个单词？"))
    num1 = int(input("你需要复习多少个单词？"))
    path = input("默写文件地址：")
    path1 = input("复习文件地址：")
    wb = openpyxl.load_workbook(path)
    wb_review = openpyxl.load_workbook(path1)

    sheets = wb.get_sheet_names()
    sheets_review = wb_review.get_sheet_names()

    data_list = []
    for i in range(len(sheets)):
        sheet = wb.get_sheet_by_name(sheets[i])

    for i in range(len(sheets_review)):
        sheet_review = wb_review.get_sheet_by_name(sheets_review[i])

        for rownum in random.sample(range(1,100), num):
            for columnnum in range(1, sheet.max_column+1):
                print(sheet.cell(row = rownum, column = columnnum).value)
                data_list.append(sheet.cell(row = rownum, column = columnnum).value)
        for rownum1 in random.sample(range(1,100), num1):
            for columnnum1 in range(1, sheet_review.max_column+1):
                data_list.append(sheet_review.cell(row = rownum1, column = columnnum1).value)

    for i in data_list:
        print(i)
        print('s1')

    a = []
    for i in  range(len(data_list)):
        if i % 2 == 0:
            a.append(data_list[i])

    for i in a:
        print(i)
        

    wbl = Workbook()

    new_ws = wbl.create_sheet(title = 'test')
    key_ws = wbl.create_sheet(title = 'test_key')
    i = 1
    k = 1
    for data_1 in range(len(data_list)):
        key_ws.cell(column=k, row=i, value=data_list[data_1])
        if k % 2 == 1:
            k = 2
        else:
            k = 1
            i += 1
    i = 1
    for data_m in range(len(a)):
        new_ws.cell(column=1, row=i, value=a[data_m])
        i += 1
    wbl.save(filename = 'new_file.xlsx')




    
